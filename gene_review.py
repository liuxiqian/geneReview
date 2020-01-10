# %%
import xlrd
import xlwt
from bs4 import BeautifulSoup
import openpyxl
import re


GHR_BASE_URL = 'https://ghr.nlm.nih.gov'
GHR_GENE_BASE_URL = 'https://www.ncbi.nlm.nih.gov/books/'

# %% page test
# NBK_id = 'NBK1363'
# test_url = 'https://www.ncbi.nlm.nih.gov/books/' + NBK_id
# headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36'}
# url = Request(test_url, headers=headers)


def read_excel():
    workbook = xlrd.open_workbook('/home/liuxi/Documents/项目/gene_review/download-bookss.xls')
    sheet_names = workbook.sheet_names()
    # print(sheet_names)
    for sheet_name in sheet_names:
        # print(sheet_name)
        if sheet_name == 'Sheet1':
            sheet_xls = workbook.sheet_by_name(sheet_name)
            col = sheet_xls.col_values(0)  # 获取第一列内容
            generate_intermediate_sheet(col)


def generate_intermediate_sheet(col):
    lists = []
    row_count = 1
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("intermediate_sheet")
    sheet.write(0, 0, '#NBK_id')
    for columns in col:
        # print(columns)
        sheet.write(row_count, 0, columns)
        # URL地址
        path = '/home/liuxi/Documents/项目/geneReview/books/' + columns + '.html'
        # headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36c'}
        # req = requests.get(test_url)
        # 抓取页面数据
        with open(path, 'r') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            key_word_array = soup.find_all(text=(u"In this GeneReview"))
            # 搜索关键字是否存在
            if key_word_array:
                key_word = key_word_array[0]
                # print(key_word)
                table = key_word.find_next('ul')
                # print(table)
                column_count = 1
                for child in table.childGenerator():
                    delimiter = '!@#$%'
                    name = child.getText(delimiter).split(delimiter)[0]
                    # print(name)
                    sheet.write(row_count, column_count, name)  # row, column, value
                    lists.append(name)
                    column_count = column_count+1
            else:
                sheet.write(row_count, 1, '')
        row_count = row_count + 1
    workbook.save('test_gene_review_page1.xls')
    # print(os.getcwd())
    # print(lists)
    new_lists = list(set(lists))
    print(new_lists)
    generate_module_data_sheet(new_lists, col)
    # generate_gene_review_data_sheet(col)


def generate_module_data_sheet(lists, col):
    workbook = openpyxl.Workbook()
    ws1 = workbook.active
    ws1.title = 'module_data'
    ws1['A1'] = '#NBK_id'
    for i in range(len(lists)):
        ws1.cell(1, i+2).value = lists[i]
    row_count = 2
    for columns in col:
        # print(columns)
        ws1.cell(row_count, 1).value = columns
        # URL地址
        path = '/home/liuxi/Documents/项目/gene_review/books/' + columns + '.html'
        # 抓取页面数据
        with open(path, 'r') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            key_word_array = soup.find_all(text=(u"In this GeneReview"))
            # 搜索关键字是否存在
            if key_word_array:
                key_word = key_word_array[0]
                # print(key_word)
                table = key_word.find_next('ul')
                # print(table)
                for child in table.childGenerator():
                    delimiter = '!@#$%'
                    name = child.getText(delimiter).split(delimiter)[0]
                    link = child.find('a')['href']
                    link_name = link.split('#')[1]
                    link_text = soup.find('div', id=link_name).getText()
                    column_count = 2
                    for i in range(len(lists)):
                        if lists[i] == name:
                            ws1.cell(row_count, column_count).value = link_text  # row, column, value
                        column_count = column_count + 1
            else:
                ws1.cell(row_count, 2).value = ''
        row_count = row_count + 1
    workbook.save('test_gene_review_page2.xls')
    generate_gene_review_data_sheet(col)


def generate_gene_review_data_sheet(col):
    workbook = openpyxl.Workbook()
    worksheet1 = workbook.active
    worksheet1.title = 'gene_review_data'
    worksheet1['A1'] = 'book'
    worksheet1['B1'] = '对应模块'
    worksheet1['C1'] = '对应模块具体内容'
    worksheet1['D1'] = '致病机制'
    row_count = 2
    for columns in col:
        print(columns)
        worksheet1.cell(row_count, 1).value = columns
        # URL地址
        path = '/home/liuxi/Documents/项目/gene_review/books/' + columns + '.html'
        # 抓取页面数据
        with open(path, 'r') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            key_word_array = soup.find_all(text=(u"In this GeneReview"))
            # 搜索关键字是否存在
            if key_word_array:
                key_word = key_word_array[0]
                # print(key_word)
                table = key_word.find_next('ul')
                # print(table)
                for child in table.childGenerator():
                    delimiter = '!@#$%'
                    name = child.getText(delimiter).split(delimiter)[0]
                    link = child.find('a')['href']
                    link_name = link.split('#')[1]
                    link_text = soup.find('div', id=link_name).getText()
                    regex = re.compile(r'Loss[-| +]of[-| +]function|Gain[-| +]of[-| +]function|activating|activation|increased +gene +dosage|dominant[-| +]negative|fusion')
                    search_objs = regex.findall(link_text)
                    search_obj = list(set(search_objs))
                    # print(search_obj)
                    if search_obj:
                        for key_value in search_obj:
                            worksheet1.cell(row_count, 4, key_value)
                            worksheet1.cell(row_count, 3, link_text)
                            worksheet1.cell(row_count, 2, name)
                            row_count = row_count + 1
    workbook.save('gene_review_page_excel.xlsx')


read_excel()
# gene_review_page_parse(cols)
