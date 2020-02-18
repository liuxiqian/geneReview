import openpyxl

# !/usr/bin/env python
# coding:utf-8

import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "phoenix_toolkit.settings")

'''
Django 版本大于等于1.7的时候，需要加上下面两句
import django
django.setup()
否则会抛出错误 django.core.exceptions.AppRegistryNotReady: Models aren't loaded yet.
'''

import django

# %% page test
# NBK_id = 'NBK1363'
# test_url = 'https://www.ncbi.nlm.nih.gov/books/' + NBK_id
# headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36'}
# url = Request(test_url, headers=headers)
excel_path = './text.xlsx'

#
if django.VERSION >= (1, 7):  # 自动判断版本
    django.setup()


def main():
    # 打开excel文件,获取工作簿对象
    wb = openpyxl.load_workbook(excel_path)
    sheets = wb.sheetnames
    # print(sheets[0], type(sheets))
    ws = wb[sheets[0]]
    from gene_review.models import Entry
    row_counts = ws.max_row
    for row_count in range(2, row_counts + 1):
        column_id = ws.cell(row_count, 1).value
        print(column_id)
        data = []
        column_counts = ws.max_column
        for column_count in range(2, column_counts + 1):
            lists = ws.cell(row_count, column_count).value
            # print(lists)
            # print(type(lists))
            row_key = ws.cell(1, column_count).value
            # print(row_key)
            # print(type(row_key))
            if lists is not None and row_key in lists:
                cup_data = (row_key, lists)
                data.append(cup_data)
                # print(data)
        dic = dict(data)
        # print(dic)
        Entry.objects.get_or_create(id=column_id, data=dic)


# 获取某个单元格的值
def get_cell_value(self, row, column):
    cell_value = self.ws.cell(row=row, column=column).value
    return cell_value


# 获取某列的所有值
def get_col_value(self, column):
    rows = self.ws.max_row
    column_data = []
    for i in range(2, rows + 1):
        cell_value = self.ws.cell(row=i, column=column).value
        column_data.append(cell_value)
    return column_data


# 获取某行所有值
def get_row_value(self, row):
    columns = self.ws.max_column
    row_data = []
    for i in range(2, columns + 1):
        cell_value = self.ws.cell(row=row, column=i).value
        row_data.append(cell_value)
    return row_data


if __name__ == "__main__":
    main()
    print('Done!')

